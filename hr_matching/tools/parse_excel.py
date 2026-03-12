"""Tool: Parse Excel files in any format into structured data."""

import json
from datetime import datetime, date
from pathlib import Path

import openpyxl


def _serialize_value(val):
    """Convert cell value to JSON-serializable type."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    if isinstance(val, (int, float, bool, str)):
        return val
    return str(val)


def _detect_header_row(sheet, max_scan=20):
    """Auto-detect the header row by finding the row with most non-empty cells."""
    best_row = 1
    best_count = 0
    for row_idx in range(1, min(max_scan + 1, sheet.max_row + 1)):
        count = sum(1 for cell in sheet[row_idx] if cell.value is not None)
        if count > best_count:
            best_count = count
            best_row = row_idx
    return best_row


def parse_excel(file_path: str, sheet_name: str = None) -> dict:
    """Parse an Excel file and return structured data.

    Args:
        file_path: Path to the Excel file (.xlsx or .xls).
        sheet_name: Optional sheet name. If not given, uses the first sheet.

    Returns:
        dict with keys: columns, row_count, sample_rows (first 5), all_data
    """
    path = Path(file_path)
    if not path.exists():
        return {"error": f"File not found: {file_path}"}

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        return {"error": f"Cannot open file: {e}"}

    if sheet_name and sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active

    header_row = _detect_header_row(sheet)

    # Extract column names
    columns = []
    for cell in sheet[header_row]:
        col_name = str(cell.value).strip() if cell.value is not None else f"col_{cell.column}"
        columns.append(col_name)

    # Extract all data rows
    all_data = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        row_values = [_serialize_value(cell.value) for cell in sheet[row_idx]]
        # Skip completely empty rows
        if all(v is None for v in row_values):
            continue
        row_dict = {}
        for i, col in enumerate(columns):
            row_dict[col] = row_values[i] if i < len(row_values) else None
        all_data.append(row_dict)

    wb.close()

    sample_rows = all_data[:5]

    return {
        "columns": columns,
        "row_count": len(all_data),
        "sample_rows": sample_rows,
        "all_data": all_data,
        "sheet_names": wb.sheetnames if hasattr(wb, 'sheetnames') else [],
    }
